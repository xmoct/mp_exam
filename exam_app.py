import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
import random
import os

class ExamApp:
    def __init__(self, root):
        self.root = root
        self.root.title("密评题库刷题软件")
        self.root.geometry("820x580")
        
        # 配色方案
        self.PRIMARY = '#4A90D9'
        self.PRIMARY_DARK = '#3A7FC8'
        self.SUCCESS = '#4CAF50'
        self.ERROR = '#E53935'
        self.BG = '#E8EEF3'
        self.CARD = '#FFFFFF'
        self.TEXT = '#2C3E50'
        self.TEXT_LIGHT = '#7F8C8D'
        self.BORDER = '#CBD5DC'
        
        # 题库数据
        self.questions = []
        self.all_questions = []
        self.current_index = 0
        self.score = 0
        self.wrong_questions = []
        self.user_answers = {}
        self.is_showing_answer = False
        self.type_filter = "全部"
        self.is_batch_mode = False
        self.batch_total = 0
        
        # 单选按钮变量
        self.single_choice_var = tk.StringVar()
        
        # 加载题库
        self.load_questions()
        
        # 创建菜单栏
        self.create_menu_bar()
        
        # 创建界面
        self.create_widgets()
        
        # 初始化导航列表
        self.init_nav_list()
        
        # 显示第一题
        self.show_question()
    
    def create_menu_bar(self):
        menu_bar = tk.Menu(self.root)
        self.root.config(menu=menu_bar)
        
        file_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="导入题库", command=self.import_questions)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)
        
        mode_menu = tk.Menu(menu_bar, tearoff=0)
        menu_bar.add_cascade(label="练习模式", menu=mode_menu)
        
        mode_menu.add_command(label="全部题目", command=lambda: self.set_filter("全部"))
        mode_menu.add_command(label="仅单选题", command=lambda: self.set_filter("单选"))
        mode_menu.add_command(label="仅多选题", command=lambda: self.set_filter("多选"))
        mode_menu.add_command(label="仅判断题", command=lambda: self.set_filter("判断"))
        
        batch_menu = tk.Menu(mode_menu, tearoff=0)
        mode_menu.add_cascade(label="批量练习", menu=batch_menu)
        batch_menu.add_command(label="30题一组", command=lambda: self.start_batch_practice(30))
        batch_menu.add_command(label="60题一组", command=lambda: self.start_batch_practice(60))
        batch_menu.add_command(label="140题一组", command=lambda: self.start_batch_practice(140))
        
        mode_menu.add_separator()
        mode_menu.add_command(label="错题练习", command=self.practice_wrong_questions)
    
    def import_questions(self):
        file_path = filedialog.askopenfilename(
            title="选择题库文件",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not file_path:
            return
        
        try:
            wb = load_workbook(file_path, data_only=False)
            sheet = wb.active
            
            imported_count = 0
            for row in sheet.iter_rows(min_row=2, values_only=False):
                if row[0] and row[0].value:
                    question = {
                        'question': str(row[0].value) if row[0].value else '',
                        'type': str(row[1].value) if row[1] and row[1].value else '',
                        'options': {},
                        'answer': str(row[10].value) if row[10] and row[10].value else '',
                        'analysis': str(row[11].value) if row[11] and row[11].value else '',
                        'chapter': str(row[12].value) if row[12] and row[12].value else '',
                        'difficulty': str(row[13].value) if row[13] and row[13].value else ''
                    }
                    
                    option_keys = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                    for i, key in enumerate(option_keys):
                        if i + 2 < len(row):
                            cell = row[i + 2]
                            if cell and cell.value is not None:
                                cell_value = cell.value
                                str_value = str(cell_value)
                                if isinstance(cell_value, int):
                                    num_str = str(cell_value)
                                    if len(num_str) == 6:
                                        part1 = num_str[:3]
                                        part2 = num_str[3:]
                                        if part1 in ['256', '512', '128', '160'] and part2 in ['256', '512', '128', '160']:
                                            str_value = f"{part1},{part2}"
                                question['options'][key] = str_value
                    
                    if question['options']:
                        self.questions.append(question)
                        self.all_questions.append(question)
                        imported_count += 1
            
            if imported_count > 0:
                self.current_index = 0
                self.score = 0
                self.user_answers = {}
                self.wrong_questions = []
                self.score_label.config(text=f"得分: {self.score}")
                self.init_nav_list()
                self.show_question()
                messagebox.showinfo("成功", f"成功导入 {imported_count} 道题目！")
            else:
                messagebox.showwarning("警告", "未能从文件中读取到有效题目！")
                
        except Exception as e:
            messagebox.showerror("错误", f"导入题库失败: {str(e)}")
    
    def set_filter(self, filter_type):
        self.type_filter = filter_type
        self.current_index = 0
        self.score = 0
        self.user_answers = {}
        self.wrong_questions = []
        self.score_label.config(text=f"得分: {self.score}")
        self.is_batch_mode = False
        
        if filter_type == "全部":
            self.questions = self.all_questions.copy()
        elif filter_type == "单选":
            self.questions = [q for q in self.all_questions if '多选' not in q.get('type', '') and '判断' not in q.get('type', '')]
        elif filter_type == "多选":
            self.questions = [q for q in self.all_questions if '多选' in q.get('type', '')]
        elif filter_type == "判断":
            self.questions = [q for q in self.all_questions if '判断' in q.get('type', '')]
        
        self.init_nav_list()
        self.show_question()
    
    def practice_wrong_questions(self):
        if not self.wrong_questions:
            messagebox.showinfo("提示", "暂无错题！")
            return
        
        self.current_index = 0
        self.score = 0
        self.user_answers = {}
        self.score_label.config(text=f"得分: {self.score}")
        self.is_batch_mode = False
        
        self.questions = [item['question'] for item in self.wrong_questions]
        
        self.init_nav_list()
        self.show_question()
        
        messagebox.showinfo("提示", f"开始错题练习，共 {len(self.questions)} 道错题")
    
    def start_batch_practice(self, count):
        if not self.all_questions:
            messagebox.showinfo("提示", "暂无题目！")
            return
        
        available_count = len(self.all_questions)
        if count > available_count:
            messagebox.showwarning("提示", f"题库中只有 {available_count} 道题目")
            count = available_count
        
        selected_indices = random.sample(range(available_count), count)
        self.questions = [self.all_questions[i] for i in selected_indices]
        
        self.current_index = 0
        self.score = 0
        self.user_answers = {}
        self.score_label.config(text=f"得分: {self.score}")
        self.is_batch_mode = True
        self.batch_total = count
        
        self.init_nav_list()
        self.show_question()
        
        messagebox.showinfo("提示", f"开始批量练习，共 {count} 道题目")
    
    def load_questions(self):
        files = ['file/01密评题库2026【1-3000】.xlsx', 'file/01密评题库2026【3001-5075】.xlsx']
        
        for file_path in files:
            if os.path.exists(file_path):
                try:
                    wb = load_workbook(file_path, data_only=False)
                    sheet = wb.active
                    
                    for row in sheet.iter_rows(min_row=2, values_only=False):
                        if row[0] and row[0].value:
                            question = {
                                'question': str(row[0].value) if row[0].value else '',
                                'type': str(row[1].value) if row[1] and row[1].value else '',
                                'options': {},
                                'answer': str(row[10].value) if row[10] and row[10].value else '',
                                'analysis': str(row[11].value) if row[11] and row[11].value else '',
                                'chapter': str(row[12].value) if row[12] and row[12].value else '',
                                'difficulty': str(row[13].value) if row[13] and row[13].value else ''
                            }
                            
                            option_keys = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
                            for i, key in enumerate(option_keys):
                                if i + 2 < len(row):
                                    cell = row[i + 2]
                                    if cell and cell.value is not None:
                                        cell_value = cell.value
                                        str_value = str(cell_value)
                                        if isinstance(cell_value, int):
                                            num_str = str(cell_value)
                                            if len(num_str) == 6:
                                                part1 = num_str[:3]
                                                part2 = num_str[3:]
                                                if part1 in ['256', '512', '128', '160'] and part2 in ['256', '512', '128', '160']:
                                                    str_value = f"{part1},{part2}"
                                        question['options'][key] = str_value
                            
                            if question['options']:
                                self.questions.append(question)
                                self.all_questions.append(question)
                except Exception as e:
                    messagebox.showwarning("警告", f"加载文件 {file_path} 失败: {str(e)}")
        
        if not self.questions:
            messagebox.showerror("错误", "未找到题库数据！")
            self.root.quit()
    
    def create_widgets(self):
        # 使用类属性的配色方案
        PRIMARY = self.PRIMARY
        PRIMARY_DARK = self.PRIMARY_DARK
        SUCCESS = self.SUCCESS
        ERROR = self.ERROR
        BG = self.BG
        CARD = self.CARD
        TEXT = self.TEXT
        TEXT_LIGHT = self.TEXT_LIGHT
        BORDER = self.BORDER
        
        self.root.configure(bg=BG)
        
        # 顶部栏
        self.header_frame = tk.Frame(self.root, bg=PRIMARY, height=48)
        self.header_frame.pack(fill=tk.X)
        self.header_frame.pack_propagate(False)
        
        header_inner = tk.Frame(self.header_frame, bg=PRIMARY)
        header_inner.pack(fill=tk.BOTH, expand=True, padx=20, pady=8)
        
        tk.Label(header_inner, text="密评题库刷题", font=("微软雅黑", 16, 'bold'), 
                 fg='white', bg=PRIMARY).pack(side=tk.LEFT)
        
        status_frame = tk.Frame(header_inner, bg=PRIMARY)
        status_frame.pack(side=tk.RIGHT)
        
        self.progress_label = tk.Label(status_frame, text="进度: 0/0", font=("微软雅黑", 12, 'bold'), 
                                       fg='white', bg=PRIMARY)
        self.progress_label.pack(side=tk.LEFT, padx=(0, 20))
        
        self.score_label = tk.Label(status_frame, text="得分: 0", font=("微软雅黑", 12, 'bold'),
                                     fg='white', bg=PRIMARY)
        self.score_label.pack(side=tk.RIGHT)
        
        # 主体区域
        self.main_frame = tk.Frame(self.root, bg=BG)
        self.main_frame.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
        
        # 左侧导航
        self.nav_frame = tk.Frame(self.main_frame, bg=CARD, width=95, relief=tk.RIDGE, bd=1)
        self.nav_frame.pack(side=tk.LEFT, fill=tk.Y, padx=(0, 10))
        self.nav_frame.pack_propagate(False)
        
        nav_inner = tk.Frame(self.nav_frame, bg=CARD)
        nav_inner.pack(fill=tk.BOTH, expand=True, padx=6, pady=8)
        
        tk.Label(nav_inner, text="题目导航", font=("微软雅黑", 11, 'bold'), bg=CARD, fg=TEXT).pack(pady=(0, 6))
        
        # 图例
        legend_frame = tk.Frame(nav_inner, bg=CARD)
        legend_frame.pack(pady=(0, 6))
        
        for symbol, color, label in [("○", TEXT_LIGHT, "未做"), ("●", SUCCESS, "正确"), ("●", ERROR, "错误")]:
            item = tk.Frame(legend_frame, bg=CARD)
            item.pack(anchor=tk.W, pady=1)
            tk.Label(item, text=symbol, fg=color, font=("微软雅黑", 9), bg=CARD).pack(side=tk.LEFT)
            tk.Label(item, text=label, font=("微软雅黑", 8), fg=TEXT_LIGHT, bg=CARD).pack(side=tk.LEFT, padx=2)
        
        # 导航列表
        self.nav_scrollbar = ttk.Scrollbar(nav_inner, orient="vertical")
        self.nav_listbox = tk.Listbox(nav_inner, font=("微软雅黑", 10), width=6, height=24,
                                   selectmode=tk.SINGLE, activestyle='none',
                                   yscrollcommand=self.nav_scrollbar.set,
                                   bg=CARD, bd=0, highlightthickness=0,
                                   selectbackground='#E8F0FE', selectforeground=PRIMARY)
        self.nav_scrollbar.config(command=self.nav_listbox.yview)
        
        self.nav_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.nav_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.nav_listbox.bind('<<ListboxSelect>>', self.on_nav_select)
        self.nav_listbox.bind("<MouseWheel>", self.on_listbox_mousewheel)
        
        # 右侧内容
        self.content_frame = tk.Frame(self.main_frame, bg=BG)
        self.content_frame.pack(side=tk.RIGHT, fill=tk.BOTH, expand=True)
        
        # 题目卡片
        self.question_card = tk.Frame(self.content_frame, bg=CARD, relief=tk.RIDGE, bd=1)
        self.question_card.pack(fill=tk.X, pady=(0, 8))
        
        question_inner = tk.Frame(self.question_card, bg=CARD)
        question_inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=12)
        
        self.question_text = tk.Text(question_inner, wrap=tk.WORD, height=5, font=("微软雅黑", 13), 
                                    bg=CARD, relief=tk.FLAT, borderwidth=0, fg=TEXT)
        self.question_text.pack(fill=tk.BOTH, expand=True)
        self.question_text.config(state=tk.DISABLED)
        
        # 选项卡片
        self.options_card = tk.Frame(self.content_frame, bg=CARD, relief=tk.RIDGE, bd=1)
        self.options_card.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        
        self.options_frame = tk.Frame(self.options_card, bg=CARD)
        self.options_frame.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        # 答案卡片
        self.answer_card = tk.Frame(self.content_frame, bg=CARD, relief=tk.RIDGE, bd=1)
        self.answer_card.pack(fill=tk.X)
        
        answer_inner = tk.Frame(self.answer_card, bg=CARD)
        answer_inner.pack(fill=tk.BOTH, expand=True, padx=15, pady=10)
        
        self.answer_text = tk.Text(answer_inner, wrap=tk.WORD, height=3, font=("微软雅黑", 12),
                                  bg=CARD, relief=tk.FLAT, borderwidth=0, fg=TEXT)
        self.answer_text.pack(fill=tk.BOTH, expand=True)
        self.answer_text.config(state=tk.DISABLED)
        
        # 按钮区域
        self.button_frame = tk.Frame(self.root, bg=BG)
        self.button_frame.pack(fill=tk.X, padx=12, pady=(0, 10))
        
        left_buttons = tk.Frame(self.button_frame, bg=BG)
        left_buttons.pack(side=tk.LEFT)
        
        style = ttk.Style()
        style.theme_use('clam')
        
        # 选项样式 - 与题目字体一致
        style.configure('Option.TRadiobutton', font=("微软雅黑", 13), background=CARD)
        style.configure('Option.TCheckbutton', font=("微软雅黑", 13), background=CARD)
        
        # 正确/错误选项样式
        style.configure('Correct.TRadiobutton', font=("微软雅黑", 13), foreground=SUCCESS, background=CARD)
        style.configure('Correct.TCheckbutton', font=("微软雅黑", 13), foreground=SUCCESS, background=CARD)
        style.configure('Wrong.TRadiobutton', font=("微软雅黑", 13), foreground=ERROR, background=CARD)
        style.configure('Wrong.TCheckbutton', font=("微软雅黑", 13), foreground=ERROR, background=CARD)
        
        # 滑块样式
        style.configure('Vertical.TScrollbar', 
                       background=BG, 
                       troughcolor=BG, 
                       bordercolor=BORDER,
                       darkcolor=PRIMARY,
                       lightcolor=PRIMARY,
                       arrowcolor=PRIMARY)
        style.map('Vertical.TScrollbar',
                  background=[('active', PRIMARY), ('pressed', PRIMARY_DARK)])
        
        style.configure('Action.TButton', font=("微软雅黑", 11), padding=6, background=PRIMARY, foreground='white')
        style.map('Action.TButton',
                  background=[('active', PRIMARY_DARK), ('pressed', PRIMARY_DARK)],
                  foreground=[('active', 'white'), ('pressed', 'white')])
        
        self.prev_btn = ttk.Button(left_buttons, text="上一题", command=self.prev_question, style='Action.TButton')
        self.prev_btn.pack(side=tk.LEFT, padx=5)
        
        self.submit_btn = ttk.Button(left_buttons, text="提交答案", command=self.submit_answer, style='Action.TButton')
        self.submit_btn.pack(side=tk.LEFT, padx=5)
        
        self.next_btn = ttk.Button(left_buttons, text="下一题", command=self.next_question, style='Action.TButton')
        self.next_btn.pack(side=tk.LEFT, padx=5)
        
        right_buttons = tk.Frame(self.button_frame, bg=BG)
        right_buttons.pack(side=tk.RIGHT)
        
        style.configure('Secondary.TButton', font=("微软雅黑", 11), padding=6, background=CARD, bordercolor=BORDER, borderwidth=1)
        style.map('Secondary.TButton',
                  background=[('active', '#F3F4F6')])
        
        self.shuffle_btn = ttk.Button(right_buttons, text="随机打乱", command=self.shuffle_questions, style='Secondary.TButton')
        self.shuffle_btn.pack(side=tk.RIGHT, padx=5)
        
        self.show_wrong_btn = ttk.Button(right_buttons, text="错题回顾", command=self.show_wrong_questions, style='Secondary.TButton')
        self.show_wrong_btn.pack(side=tk.RIGHT, padx=5)
    
    def is_multiple_choice(self):
        question = self.questions[self.current_index]
        q_type = question['type'] if question['type'] else ''
        return '多选' in q_type or 'multiple' in q_type.lower()
    
    def is_true_false(self):
        question = self.questions[self.current_index]
        q_type = question['type'] if question['type'] else ''
        return '判断' in q_type or 'True/False' in q_type or 'true/false' in q_type.lower()
    
    def init_nav_list(self):
        self.nav_listbox.delete(0, tk.END)
        for i in range(len(self.questions)):
            self.nav_listbox.insert(tk.END, f"{i+1}")
    
    def update_nav_status(self):
        for i in range(len(self.questions)):
            if str(i) in self.user_answers:
                question = self.questions[i]
                is_tf = '判断' in (question['type'] or '')
                if is_tf:
                    correct_answer = self.get_tf_answer_key_by_index(i)
                else:
                    correct_answer = str(question['answer']).strip() if question['answer'] else ''
                    correct_answer = ''.join(correct_answer.replace(',', '').replace(' ', '').upper())
                
                user_answer = self.user_answers[str(i)]
                if '判断' in (question['type'] or ''):
                    user_answer_clean = user_answer
                elif '多选' in (question['type'] or ''):
                    user_answer_clean = ''.join(sorted(user_answer))
                else:
                    user_answer_clean = user_answer
                
                if user_answer_clean == correct_answer:
                    self.nav_listbox.itemconfig(i, fg='#52C41A')
                else:
                    self.nav_listbox.itemconfig(i, fg='#F5222D')
            else:
                self.nav_listbox.itemconfig(i, fg='#6B7280')
        
        self.nav_listbox.selection_clear(0, tk.END)
        self.nav_listbox.selection_set(self.current_index)
        self.nav_listbox.see(self.current_index)
    
    def get_tf_answer_key_by_index(self, index):
        question = self.questions[index]
        correct_answer = str(question['answer']).strip() if question['answer'] else ''
        
        true_values = ['正确', '对', 'T', 'TRUE', 'True', 'true']
        false_values = ['错误', '错', 'F', 'FALSE', 'False', 'false']
        
        for key, value in question['options'].items():
            if str(value).strip() in true_values and correct_answer in true_values:
                return key
            if str(value).strip() in false_values and correct_answer in false_values:
                return key
        
        return correct_answer
    
    def on_nav_select(self, event):
        selection = self.nav_listbox.curselection()
        if selection:
            self.current_index = selection[0]
            self.show_question()
    
    def on_listbox_mousewheel(self, event):
        self.nav_listbox.yview_scroll(int(-1 * (event.delta / 120)), "units")
    
    def get_true_false_answer_key(self):
        question = self.questions[self.current_index]
        correct_answer = str(question['answer']).strip() if question['answer'] else ''
        
        true_values = ['正确', '对', 'T', 'TRUE', 'True', 'true']
        false_values = ['错误', '错', 'F', 'FALSE', 'False', 'false']
        
        for key, value in question['options'].items():
            if str(value).strip() in true_values and correct_answer in true_values:
                return key
            if str(value).strip() in false_values and correct_answer in false_values:
                return key
        
        return correct_answer
    
    def show_question(self):
        if not self.questions:
            return
        
        question = self.questions[self.current_index]
        self.is_showing_answer = False
        
        if self.is_batch_mode:
            answered_count = len(self.user_answers)
            self.progress_label.config(text=f"进度: {self.current_index + 1}/{len(self.questions)} | 已练习: {answered_count} 题")
        else:
            self.progress_label.config(text=f"进度: {self.current_index + 1}/{len(self.questions)}")
        
        self.question_text.config(state=tk.NORMAL)
        self.question_text.delete(1.0, tk.END)
        question_info = f"【{question['type']}】\n\n{question['question']}"
        if question['chapter']:
            question_info += f"\n\n章节: {question['chapter']}"
        if question['difficulty']:
            question_info += f" | 难度: {question['difficulty']}"
        
        if self.is_multiple_choice():
            question_info += "\n\n【多选题】请选择所有正确答案"
        
        self.question_text.insert(tk.END, question_info)
        self.question_text.config(state=tk.DISABLED)
        
        for widget in self.options_frame.winfo_children():
            widget.destroy()
        
        is_multi = self.is_multiple_choice()
        if not is_multi:
            self.single_choice_var.set('__NONE__')
        
        for key, value in question['options'].items():
            opt_frame = tk.Frame(self.options_frame, bg=self.CARD)
            opt_frame.pack(anchor=tk.W, pady=6, padx=3)
            
            if is_multi:
                var = tk.IntVar()
                var.set(0)
                btn = tk.Checkbutton(opt_frame, variable=var, onvalue=1, offvalue=0,
                                    bg=self.CARD, fg=self.TEXT)
            else:
                var = self.single_choice_var
                btn = tk.Radiobutton(opt_frame, variable=var, value=key,
                                     bg=self.CARD, fg=self.TEXT)
            
            btn.pack(side=tk.LEFT)
            
            label = tk.Label(opt_frame, text=f"{key}. {value}", font=("微软雅黑", 13), 
                            bg=self.CARD, fg=self.TEXT, anchor=tk.W, cursor='hand2')
            label.pack(side=tk.LEFT, padx=5)
            label.bind('<Button-1>', lambda event, k=key: self.on_option_label_click(k))
            
            self.options_frame.__setattr__(f"opt_var_{key}", var)
            self.options_frame.__setattr__(f"opt_btn_{key}", btn)
            self.options_frame.__setattr__(f"opt_label_{key}", label)
            self.options_frame.__setattr__(f"opt_frame_{key}", opt_frame)
            self.options_frame.__setattr__(f"opt_type_{key}", 'check' if is_multi else 'radio')
        
        if str(self.current_index) in self.user_answers:
            saved_answer = self.user_answers[str(self.current_index)]
            if is_multi:
                for key in question['options'].keys():
                    var = getattr(self.options_frame, f"opt_var_{key}", None)
                    if var:
                        var.set(1 if key in saved_answer else 0)
            else:
                if saved_answer in question['options']:
                    self.single_choice_var.set(saved_answer)
                else:
                    self.single_choice_var.set('')
        
        self.update_answer_display()
        self.update_nav_status()
    
    def submit_answer(self):
        if self.is_showing_answer:
            self.next_question()
            return
        
        question = self.questions[self.current_index]
        is_multi = self.is_multiple_choice()
        is_tf = self.is_true_false()
        user_answer = ""
        
        if is_multi:
            selected = []
            for key in question['options'].keys():
                var = getattr(self.options_frame, f"opt_var_{key}", None)
                if var and var.get() == 1:
                    selected.append(key)
            user_answer = ''.join(sorted(selected))
        else:
            user_answer = getattr(self, 'single_choice_var', None)
            if user_answer:
                user_answer = user_answer.get()
        
        if not user_answer or user_answer == '__NONE__':
            messagebox.showwarning("提示", "请选择答案！")
            return
        
        self.user_answers[str(self.current_index)] = user_answer
        
        if is_tf:
            correct_answer_clean = self.get_true_false_answer_key()
        else:
            correct_answer = str(question['answer']).strip() if question['answer'] else ''
            correct_answer_clean = ''.join(correct_answer.replace(',', '').replace(' ', '').upper())
        
        if is_multi:
            user_answer_clean = ''.join(sorted(user_answer))
        else:
            user_answer_clean = user_answer
        
        correct = user_answer_clean == correct_answer_clean
        
        if correct:
            self.score += 1
            self.score_label.config(text=f"得分: {self.score}")
        else:
            if self.current_index not in [q['index'] for q in self.wrong_questions]:
                self.wrong_questions.append({
                    'index': self.current_index,
                    'question': question,
                    'user_answer': user_answer
                })
        
        self.show_answer_manual(correct)
        
        if correct and self.current_index < len(self.questions) - 1:
            self.root.after(1000, self.next_question)
        else:
            self.submit_btn.config(text="下一题")
            self.is_showing_answer = True
    
    def on_single_choice_select(self):
        if not self.is_showing_answer:
            self.submit_answer()
    
    def on_option_label_click(self, key):
        if self.is_showing_answer:
            return
        
        is_multi = self.is_multiple_choice()
        
        if is_multi:
            var = getattr(self.options_frame, f"opt_var_{key}", None)
            if var:
                var.set(1 if var.get() == 0 else 0)
        else:
            self.single_choice_var.set(key)
            self.submit_answer()
    
    def show_answer_manual(self, correct):
        question = self.questions[self.current_index]
        is_multi = self.is_multiple_choice()
        is_tf = self.is_true_false()
        
        if is_tf:
            correct_answer_key = self.get_true_false_answer_key()
            correct_answer_text = question['options'].get(correct_answer_key, str(question['answer']))
            display_answer = correct_answer_text
        else:
            correct_answer = str(question['answer']).strip() if question['answer'] else ''
            display_answer = correct_answer
        
        self.answer_text.config(state=tk.NORMAL)
        self.answer_text.delete(1.0, tk.END)
        
        if correct:
            result = "✓ 回答正确！"
            self.answer_text.insert(tk.END, result)
            self.answer_text.tag_add("correct", "1.0", "1.end")
            self.answer_text.tag_config("correct", foreground="#52C41A", font=("微软雅黑", 13, 'bold'))
        else:
            result = f"✗ 回答错误！\n正确答案: {display_answer}"
            self.answer_text.insert(tk.END, result)
            self.answer_text.tag_add("wrong", "1.0", "1.end")
            self.answer_text.tag_config("wrong", foreground="#F5222D", font=("微软雅黑", 13, 'bold'))
        
        if question['analysis']:
            self.answer_text.insert(tk.END, f"\n\n解析: {question['analysis']}")
        
        self.answer_text.config(state=tk.DISABLED)
        
        if is_tf:
            correct_keys = [self.get_true_false_answer_key()]
        else:
            correct_answer = str(question['answer']).strip() if question['answer'] else ''
            correct_keys = list(''.join(correct_answer.replace(',', '').replace(' ', '')))
        
        for key in question['options'].keys():
            label = getattr(self.options_frame, f"opt_label_{key}", None)
            if label:
                is_correct_option = key in correct_keys
                
                if is_correct_option:
                    label.config(fg='#4CAF50')
                elif not correct:
                    var = getattr(self.options_frame, f"opt_var_{key}", None)
                    if var and ((is_multi and var.get() == 1) or (not is_multi and var.get() == key)):
                        label.config(fg='#E53935')
    
    def update_answer_display(self):
        self.answer_text.config(state=tk.NORMAL)
        self.answer_text.delete(1.0, tk.END)
        
        if str(self.current_index) in self.user_answers:
            question = self.questions[self.current_index]
            user_answer = self.user_answers[str(self.current_index)]
            is_multi = self.is_multiple_choice()
            is_tf = self.is_true_false()
            
            if is_tf:
                correct_answer_clean = self.get_true_false_answer_key()
            else:
                correct_answer = str(question['answer']).strip() if question['answer'] else ''
                correct_answer_clean = ''.join(correct_answer.replace(',', '').replace(' ', '').upper())
            
            if is_multi:
                user_answer_clean = ''.join(sorted(user_answer))
            else:
                user_answer_clean = user_answer
            
            correct = user_answer_clean == correct_answer_clean
            
            self.show_answer_manual(correct)
            self.submit_btn.config(text="下一题")
            self.is_showing_answer = True
        else:
            self.answer_text.insert(tk.END, "请选择答案并点击提交")
            self.submit_btn.config(text="提交答案")
            self.is_showing_answer = False
        
        self.answer_text.config(state=tk.DISABLED)
        
        question = self.questions[self.current_index]
        
        for key in question['options'].keys():
            label = getattr(self.options_frame, f"opt_label_{key}", None)
            if label:
                label.config(fg=self.TEXT)
    
    def prev_question(self):
        if self.current_index > 0:
            self.current_index -= 1
            self.show_question()
    
    def next_question(self):
        if self.current_index < len(self.questions) - 1:
            self.current_index += 1
            self.show_question()
        else:
            self.show_results()
    
    def shuffle_questions(self):
        random.shuffle(self.questions)
        self.current_index = 0
        self.score = 0
        self.user_answers = {}
        self.wrong_questions = []
        self.score_label.config(text=f"得分: {self.score}")
        self.init_nav_list()
        self.show_question()
        messagebox.showinfo("提示", "题目已随机打乱！")
    
    def show_wrong_questions(self):
        if not self.wrong_questions:
            messagebox.showinfo("提示", "暂无错题！")
            return
        
        wrong_window = tk.Toplevel(self.root)
        wrong_window.title("错题回顾")
        wrong_window.geometry("780x550")
        wrong_window.configure(bg='#F4F6F8')
        
        listbox = tk.Listbox(wrong_window, font=("微软雅黑", 11), width=85, height=18, bg='#FFFFFF')
        listbox.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        for i, item in enumerate(self.wrong_questions):
            q = item['question']
            listbox.insert(tk.END, f"{i+1}. {q['question'][:50]}...")
        
        def view_detail():
            selection = listbox.curselection()
            if selection:
                idx = selection[0]
                item = self.wrong_questions[idx]
                q = item['question']
                
                detail_window = tk.Toplevel(wrong_window)
                detail_window.title("错题详情")
                detail_window.geometry("780x600")
                detail_window.configure(bg='#F4F6F8')
                
                q_frame = tk.Frame(detail_window, bg='#FFFFFF')
                q_frame.pack(fill=tk.X, padx=10, pady=8)
                
                q_text = tk.Text(q_frame, wrap=tk.WORD, height=7, font=("微软雅黑", 12),
                               bg='#FFFFFF', relief=tk.FLAT)
                q_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                q_text.insert(tk.END, f"【{q['type']}】\n\n{q['question']}")
                q_text.config(state=tk.DISABLED)
                
                opt_frame = tk.Frame(detail_window, bg='#FFFFFF')
                opt_frame.pack(fill=tk.X, padx=10, pady=8)
                
                is_tf = '判断' in (q['type'] or '')
                if is_tf:
                    true_values = ['正确', '对', 'T', 'TRUE', 'True', 'true']
                    false_values = ['错误', '错', 'F', 'FALSE', 'False', 'false']
                    correct_answer = str(q['answer']).strip() if q['answer'] else ''
                    correct_key = None
                    for key, value in q['options'].items():
                        if str(value).strip() in true_values and correct_answer in true_values:
                            correct_key = key
                            break
                        if str(value).strip() in false_values and correct_answer in false_values:
                            correct_key = key
                            break
                    correct_answer_clean = correct_key if correct_key else correct_answer
                else:
                    correct_answer = str(q['answer']).strip() if q['answer'] else ''
                    correct_answer_clean = ''.join(correct_answer.replace(',', '').replace(' ', ''))
                
                for key, value in q['options'].items():
                    is_correct = key in str(correct_answer_clean)
                    is_user = key in str(item['user_answer'])
                    
                    if is_correct:
                        color = "#52C41A"
                    elif is_user:
                        color = "#F5222D"
                    else:
                        color = "#374151"
                    
                    label = tk.Label(opt_frame, text=f"{key}. {value}", fg=color, font=("微软雅黑", 12), bg='#FFFFFF')
                    label.pack(anchor=tk.W, pady=5, padx=12)
                
                ans_frame = tk.Frame(detail_window, bg='#FFFFFF')
                ans_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=8)
                
                ans_text = tk.Text(ans_frame, wrap=tk.WORD, height=6, font=("微软雅黑", 12),
                                  bg='#F9FAFB')
                ans_text.pack(fill=tk.BOTH, expand=True, padx=12, pady=10)
                
                ans_text.insert(tk.END, f"你的答案: {item['user_answer']}\n")
                ans_text.insert(tk.END, f"正确答案: {correct_answer}\n")
                if q['analysis']:
                    ans_text.insert(tk.END, f"\n解析: {q['analysis']}")
                ans_text.config(state=tk.DISABLED)
        
        ttk.Button(wrong_window, text="查看详情", command=view_detail).pack(pady=10)
    
    def show_results(self):
        total = len(self.questions)
        correct = self.score
        wrong = total - correct
        accuracy = (correct / total) * 100
        
        result_text = f"""
测试完成！

总题数: {total}
正确: {correct}
错误: {wrong}
正确率: {accuracy:.2f}%
"""
        
        messagebox.showinfo("测试结果", result_text)
        
        if messagebox.askyesno("重新开始", "是否重新开始刷题？"):
            self.shuffle_questions()
        else:
            self.root.quit()

if __name__ == "__main__":
    root = tk.Tk()
    app = ExamApp(root)
    root.mainloop()